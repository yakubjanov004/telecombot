import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Generator, List, Optional, Union

import openpyxl

from backend.utils.logger import logger

COLUMN_ALIASES = {
    "BRANCHES": [
        "BRANCH",
        "BRANCHES",
        "ФИЛИАЛ",
        "ФИЛИАЛЫ",
        "HUDUD",
        "REGION",
        "РЕГИОН",
        "РЕГИОНЫ",
    ],
    "NAVI_USER": [
        "NAVI_USER",
        "NAVI_USERNAME",
        "USER",
        "OPERATOR",
        "NAVI USER",
        "NAVIUSER",
        "NAVIUSERNAME",
    ],
    "RT_LC_STATES": ["RT_LC_STATES", "STATUS", "STATE", "ACTIVE STATUS", "ST", "СОСТОЯНИЕ"],
    "MSISDN": ["MSISDN", "PHONE", "NUMBER", "RAQAM", "MOBILE_NUMBER", "АБОНЕНТСКИЙ НОМЕР"],
    "RATE_PLAN": ["RATE_PLAN", "RP", "TARIF", "RATE_PLAN_FIRST_CONNECTION", "ТАРИФНЫЙ ПЛАН"],
    "AMOUNT": [
        "HAQI",
        "ПРОДАЖИ",
        "НАЧИСЛЕНИЯ",
        "НАЧИСЛЕНО",
        "AMOUNT",
        "SUMMA",
        "CHARGED AMOUNT",
        "СУММА",
        "ПРОДАЖИ, НАЧИСЛЕНИЯ (НДС)",
        "НАЧИСЛЕНО АП",
    ],
    "QUANTITY": ["QUANTITY", "SONI", "COUNT", "QTY", "КОЛИЧЕСТВО", "ПРОДАЖИ, К-ВО"],
    "ACTIVATION_DATE": [
        "ACTIVATION DATE",
        "SANASI",
        "DATE",
        "ACT_DATE",
        "ДАТА АКТИВАЦИИ",
        "ACTIVATION_DATE",
    ],
    "DEALER": ["DEALER", "DILLER", "AGENT", "ДИЛЕР"],
    "SALE_POINT": ["SALE POINT", "SALE_POINT", "TOCHKA", "ТТ", "TT", "ТОЧКА ПРОДАЖИ"],
    "STANDARD_TYPE": ["STANDARD TYPE", "STANDARD", "TECH", "TYPE", "СТАНДАРТ", "STANDARTS"],
    "DEPARTMENTS": ["DEPARTMENTS", "DEPARTMENT", "ОТДЕЛ", "РУТ", "ГУТ", "ПОДРАЗДЕЛЕНИЕ"],
}

MONTH_MAP = {
    "yanvar": 1,
    "yan": 1,
    "january": 1,
    "jan": 1,
    "январь": 1,
    "январ": 1,
    "янв": 1,
    "fevral": 2,
    "fev": 2,
    "february": 2,
    "feb": 2,
    "февраль": 2,
    "феврал": 2,
    "фев": 2,
    "mart": 3,
    "mar": 3,
    "march": 3,
    "март": 3,
    "мар": 3,
    "aprel": 4,
    "apr": 4,
    "april": 4,
    "апрель": 4,
    "апрел": 4,
    "апр": 4,
    "may": 5,
    "май": 5,
    "iyun": 6,
    "jun": 6,
    "june": 6,
    "июнь": 6,
    "июн": 6,
    "iyul": 7,
    "jul": 7,
    "july": 7,
    "июль": 7,
    "июл": 7,
    "avgust": 8,
    "avg": 8,
    "august": 8,
    "aug": 8,
    "август": 8,
    "авг": 8,
    "sentabr": 9,
    "sep": 9,
    "september": 9,
    "сентябрь": 9,
    "сентябр": 9,
    "сент": 9,
    "oktabr": 10,
    "oct": 10,
    "october": 10,
    "октябрь": 10,
    "октябр": 10,
    "окт": 10,
    "noyabr": 11,
    "nov": 11,
    "november": 11,
    "ноябрь": 11,
    "ноябр": 11,
    "ноя": 11,
    "dekabr": 12,
    "dec": 12,
    "december": 12,
    "декабрь": 12,
    "декабр": 12,
    "дек": 12,
}


class ExcelReader:
    def __init__(self, file_path: Optional[Union[str, Path]] = None):
        if file_path:
            self.file_path = Path(file_path)
            if not self.file_path.exists():
                logger.warning(f"Excel file not found: {file_path}")
        else:
            self.file_path = None

    def get_sheets(self) -> List[str]:
        if not self.file_path or not self.file_path.exists():
            return []
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception as e:
            logger.error(f"Error reading sheet names from {self.file_path}: {e}")
            return []

    def parse_date(self, value: Any) -> Optional[date]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            try:
                return date(1899, 12, 30) + timedelta(days=int(value))
            except Exception:
                return None
        if isinstance(value, str):
            value = value.strip()
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return None

    def extract_period(self, text: str) -> Dict[str, Optional[int]]:
        text_lower = str(text).lower().strip()
        text_lower = re.sub(r"(\.xlsx)?\.dec$", "", text_lower)
        text_lower = re.sub(r"\.xlsx$", "", text_lower)

        found_month = None
        tokens = re.findall(r"[a-zа-яё]+", text_lower, flags=re.IGNORECASE)
        for token in tokens:
            month = MONTH_MAP.get(token)
            if month:
                found_month = month
                break

        year_match = re.search(r"20[0-9]{2}", str(text))
        found_year = int(year_match.group()) if year_match else None

        return {"month": found_month, "year": found_year}

    def map_header(self, raw_header: str) -> str:
        clean = str(raw_header).strip().upper()
        clean = " ".join(clean.split())
        clean_underscore = clean.replace(" ", "_")
        for standard, aliases in COLUMN_ALIASES.items():
            if clean in aliases or clean_underscore in aliases:
                return standard
        return clean

    def read_sheet(self, sheet_name: str) -> Generator[Dict[str, Any], None, None]:
        if not self.file_path:
            return

        wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)
        sheet = wb[sheet_name]

        header_row = None
        header_row_number = 1
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=5, values_only=True),
            start=1,
        ):
            if any(row):
                header_row = row
                header_row_number = row_number
                break

        if not header_row:
            wb.close()
            return

        mapped_headers = [
            self.map_header(h) if h else f"UNKNOWN_{i}" for i, h in enumerate(header_row)
        ]

        for row_idx, row_values in enumerate(
            sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            if not any(row_values):
                continue
            row_data = {
                mapped_headers[i]: value
                for i, value in enumerate(row_values)
                if i < len(mapped_headers)
            }
            row_data["ROW_NUMBER"] = row_idx
            if "ACTIVATION_DATE" in row_data:
                row_data["ACTIVATION_DATE"] = self.parse_date(row_data["ACTIVATION_DATE"])
            yield row_data

        wb.close()


def detect_service_type_by_headers(headers: List[str]) -> str:
    h_set = {str(h).upper() for h in headers if h}

    if "DEPARTMENTS" in h_set and "RT_LC_STATES" in h_set:
        return "internet"

    if (
        "DEALER" in h_set
        and "BRANCHES" in h_set
        and "MSISDN" in h_set
        and "DEPARTMENTS" not in h_set
        and "RT_LC_STATES" not in h_set
    ):
        return "mobile"

    if "RT_LC_STATES" in h_set:
        return "internet"
    if "SALE_POINT" in h_set:
        return "mobile"
    if "STANDARD_TYPE" in h_set:
        return "internet"
    if "DEALER" in h_set and "MSISDN" in h_set and "RT_LC_STATES" not in h_set:
        return "mobile"

    return "unknown"
