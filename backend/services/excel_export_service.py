import copy
import os
import re
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.models import (
    Branch,
    Dealer,
    ImportBatch,
    InternetSale,
    MobileSale,
    RatePlan,
    SalePoint,
    SaleSource,
    ServiceType,
)
from backend.utils.excel_reader import ExcelReader


PROJECT_ROOT = Path(__file__).resolve().parents[2]
EXCEL_ROOT = PROJECT_ROOT / "assets" / "excel"


def _discover_excel_dirs() -> list[Path]:
    if not EXCEL_ROOT.exists():
        return []

    directories = [
        p
        for p in sorted(EXCEL_ROOT.iterdir(), key=lambda p: p.name.lower())
        if p.is_dir() and any(f.suffix.lower() == ".xlsx" for f in p.iterdir())
    ]

    if directories:
        return directories

    return []


EXCEL_DIRS = _discover_excel_dirs()

DEFAULT_INTERNET_HEADERS = [
    "BRANCHES",
    "DEPARTMENTS",
    "NAVI_USER",
    "RT_LC_STATES",
    "MSISDN",
    "RATE_PLAN_FIRST_CONNECTION",
]
DEFAULT_MOBILE_HEADERS = [
    "DEALER",
    "NAVI_USER",
    "MSISDN",
    "RATE_PLAN_FIRST_CONNECTION",
    "Branches",
]


def _clean_sheet_title(title: str, used_titles: set[str]) -> str:
    clean = re.sub(r"[:\\/?*\[\]]", " ", title).strip() or "Sheet"
    clean = clean[:31]
    if clean not in used_titles:
        used_titles.add(clean)
        return clean

    base = clean[:28].rstrip()
    index = 2
    while True:
        candidate = f"{base} {index}"[:31]
        if candidate not in used_titles:
            used_titles.add(candidate)
            return candidate
        index += 1


def _source_base_name(file_name: str) -> str:
    if file_name.endswith(".dec.xlsx"):
        file_name = file_name.removesuffix(".dec.xlsx")
    if file_name.endswith(".xlsx"):
        file_name = file_name.removesuffix(".xlsx")
    return file_name


def _find_source_file(file_name: str | None) -> Path | None:
    if not file_name:
        return None

    base = _source_base_name(file_name)
    candidates = [
        f"{base}.dec.xlsx",
        f"{base}.xlsx.dec.xlsx",
        f"{base}.xlsx",
        file_name,
    ]

    for data_dir in EXCEL_DIRS:
        for candidate in candidates:
            path = data_dir / candidate
            if path.exists():
                return path
    return None


def _first_header_row(ws) -> tuple[int, list[Any]]:
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=5), start=1):
        values = [cell.value for cell in row]
        if any(value is not None for value in values):
            return row_number, values
    return 1, []


def _default_headers(batch: ImportBatch) -> list[str]:
    if batch.service_type == ServiceType.INTERNET:
        return DEFAULT_INTERNET_HEADERS
    return DEFAULT_MOBILE_HEADERS


def _read_source_layout(batch: ImportBatch):
    source_file = _find_source_file(batch.file_name)
    if not source_file:
        return None, 1, _default_headers(batch)

    wb = openpyxl.load_workbook(source_file, read_only=False, data_only=True)
    try:
        if batch.sheet_name not in wb.sheetnames:
            return None, 1, _default_headers(batch)
        source_ws = wb[batch.sheet_name]
        header_row_num, headers = _first_header_row(source_ws)
        return source_ws, header_row_num, headers
    except Exception:
        wb.close()
        raise


def _apply_source_header_layout(target_ws, source_ws, header_row_num: int, header_count: int) -> None:
    if not source_ws:
        return

    target_ws.row_dimensions[1].height = source_ws.row_dimensions[header_row_num].height
    for col_idx in range(1, header_count + 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        target_ws.column_dimensions[letter].width = source_ws.column_dimensions[letter].width
        source_cell = source_ws.cell(header_row_num, col_idx)
        target_cell = target_ws.cell(1, col_idx)
        if source_cell.has_style:
            target_cell.font = copy.copy(source_cell.font)
            target_cell.fill = copy.copy(source_cell.fill)
            target_cell.border = copy.copy(source_cell.border)
            target_cell.alignment = copy.copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy.copy(source_cell.protection)


def _batch_sheet_title(batch: ImportBatch) -> str:
    base = _source_base_name(batch.file_name or "")
    if base:
        return f"{base} {batch.sheet_name or ''}".strip()
    prefix = "Internet" if batch.service_type == ServiceType.INTERNET else "Mobile"
    return f"{prefix} {batch.sheet_name or batch.id}"


def _value_for_header(
    reader: ExcelReader,
    header: Any,
    sale: InternetSale | MobileSale,
    branch_name: str | None,
    dealer_name: str | None,
    rate_plan_name: str | None,
    sale_point_name: str | None = None,
) -> Any:
    standard = reader.map_header(header)

    if standard == "BRANCHES":
        return branch_name or sale.branch_name_raw or ""
    if standard == "DEALER":
        return dealer_name or getattr(sale, "dealer_name_raw", None) or ""
    if standard == "SALE_POINT":
        return sale_point_name or getattr(sale, "sale_point_name_raw", None) or ""
    if standard == "DEPARTMENTS":
        return getattr(sale, "department_name_raw", None) or ""
    if standard == "STANDARD_TYPE":
        return getattr(sale, "standard_type", None) or ""
    if standard == "NAVI_USER":
        return sale.navi_user or ""
    if standard == "RT_LC_STATES":
        return getattr(sale, "rt_lc_state", None) or ""
    if standard == "MSISDN":
        return sale.msisdn or ""
    if standard == "RATE_PLAN":
        return rate_plan_name or sale.rate_plan_raw or ""
    if standard == "AMOUNT":
        if isinstance(sale, InternetSale):
            return sale.sale_amount or 0
        return sale.charged_amount or 0
    if standard == "QUANTITY":
        return sale.sale_quantity
    if standard == "ACTIVATION_DATE":
        return sale.activation_date

    return ""


async def _internet_rows(session: AsyncSession, period_id: int, batch_id: int | None):
    stmt = (
        select(
            InternetSale,
            Branch.name.label("branch_name"),
            Dealer.name.label("dealer_name"),
            RatePlan.name.label("rate_plan_name"),
        )
        .outerjoin(Branch, InternetSale.branch_id == Branch.id)
        .outerjoin(Dealer, InternetSale.dealer_id == Dealer.id)
        .outerjoin(RatePlan, InternetSale.rate_plan_id == RatePlan.id)
        .where(InternetSale.period_id == period_id)
        .order_by(InternetSale.row_number, InternetSale.id)
    )
    if batch_id is None:
        stmt = stmt.where(
            InternetSale.import_batch_id.is_(None),
            InternetSale.source == SaleSource.MANUAL,
        )
    else:
        stmt = stmt.where(InternetSale.import_batch_id == batch_id)

    return (await session.execute(stmt)).all()


async def _mobile_rows(session: AsyncSession, period_id: int, batch_id: int | None):
    stmt = (
        select(
            MobileSale,
            Branch.name.label("branch_name"),
            Dealer.name.label("dealer_name"),
            RatePlan.name.label("rate_plan_name"),
            SalePoint.name.label("sale_point_name"),
        )
        .outerjoin(Branch, MobileSale.branch_id == Branch.id)
        .outerjoin(Dealer, MobileSale.dealer_id == Dealer.id)
        .outerjoin(RatePlan, MobileSale.rate_plan_id == RatePlan.id)
        .outerjoin(SalePoint, MobileSale.sale_point_id == SalePoint.id)
        .where(MobileSale.period_id == period_id)
        .order_by(MobileSale.row_number, MobileSale.id)
    )
    if batch_id is None:
        stmt = stmt.where(
            MobileSale.import_batch_id.is_(None),
            MobileSale.source == SaleSource.MANUAL,
        )
    else:
        stmt = stmt.where(MobileSale.import_batch_id == batch_id)

    return (await session.execute(stmt)).all()


def _append_rows(ws, reader: ExcelReader, headers: list[Any], rows: list, service_type: ServiceType) -> None:
    if service_type == ServiceType.INTERNET:
        for sale, branch_name, dealer_name, rate_plan_name in rows:
            ws.append([
                _value_for_header(
                    reader,
                    header,
                    sale,
                    branch_name,
                    dealer_name,
                    rate_plan_name,
                )
                for header in headers
            ])
    else:
        for sale, branch_name, dealer_name, rate_plan_name, sale_point_name in rows:
            ws.append([
                _value_for_header(
                    reader,
                    header,
                    sale,
                    branch_name,
                    dealer_name,
                    rate_plan_name,
                    sale_point_name,
                )
                for header in headers
            ])


async def build_period_export(session: AsyncSession, period_id: int) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()
    reader = ExcelReader()

    batches = (
        await session.execute(
            select(ImportBatch)
            .where(ImportBatch.period_id == period_id)
            .order_by(ImportBatch.id)
        )
    ).scalars().all()

    for batch in batches:
        if batch.service_type == ServiceType.INTERNET:
            rows = await _internet_rows(session, period_id, batch.id)
        else:
            rows = await _mobile_rows(session, period_id, batch.id)
        if not rows:
            continue

        source_ws, header_row_num, headers = _read_source_layout(batch)
        headers = [header for header in headers if header is not None]
        if not headers:
            headers = _default_headers(batch)

        ws = wb.create_sheet(_clean_sheet_title(_batch_sheet_title(batch), used_titles))
        ws.append(headers)
        _apply_source_header_layout(ws, source_ws, header_row_num, len(headers))
        if source_ws:
            source_ws.parent.close()
        _append_rows(ws, reader, headers, rows, batch.service_type)

    manual_internet = await _internet_rows(session, period_id, None)
    if manual_internet:
        ws = wb.create_sheet(_clean_sheet_title("Manual Internet", used_titles))
        ws.append(DEFAULT_INTERNET_HEADERS)
        _append_rows(ws, reader, DEFAULT_INTERNET_HEADERS, manual_internet, ServiceType.INTERNET)

    manual_mobile = await _mobile_rows(session, period_id, None)
    if manual_mobile:
        ws = wb.create_sheet(_clean_sheet_title("Manual Mobile", used_titles))
        ws.append(DEFAULT_MOBILE_HEADERS)
        _append_rows(ws, reader, DEFAULT_MOBILE_HEADERS, manual_mobile, ServiceType.MOBILE)

    if not wb.worksheets:
        ws = wb.create_sheet("Empty")
        ws.append(["No data"])

    output_dir = PROJECT_ROOT / "storage" / "exports"
    os.makedirs(output_dir, exist_ok=True)
    export_path = output_dir / f"export_period_{period_id}.xlsx"
    wb.save(export_path)
    return str(export_path)
